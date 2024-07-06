import csv
import openpyxl
import pyautogui
import datetime
import os
import time

def read_file_path1(filename):
    with open(filename, "r") as file:
        lines = file.readlines()

    file_path_1 = lines[2].strip()

    return file_path_1

def read_file_path2(filename):
    with open(filename, "r") as file:
        lines = file.readlines()

    file_path_2 = lines[1].strip()
    

    return file_path_2

def read_file_path3(filename):
    with open(filename, "r") as file:
        lines = file.readlines()

    file_path_3 = lines[0].strip()

    return file_path_3
path0 = read_file_path3("C:\Attendance\Code\_Important.txt")
path1 = (path0 + r"\attendance")
path2 = (path0 + "\Data_From_attendance.txt")
path3 = (path0 + "\Mac_Address_Before.txt")
path4 = (path0 + "\Pure_Mac_Address.txt")
path5 = (path0 + "\\Unique_Mac_Address.txt")

coding1 = "utf-16 LE"
coding2 = "utf-8"

f = open(path1, "r", encoding=coding1)
content = f.read()
f.close()
f = open(path2, "w", encoding=coding2)
f.write(content)
f.close()
f = open(path3, "w", encoding=coding2)
f.write("")
f.close()

x = open(path2,'r')
for line in x :
    if line.startswith("MAC Address"):
        open(path3, "a" ,encoding=coding2).write(line)


n = open(path3, "r")
g = open(path4, "a")

for line in n:
    if line.strip():
        g.write("\t".join(line.split()[3:]) + "\n")

n.close()
g.close()
# Read the MAC addresses from the input file
with open(path4, 'r') as input_file:
    mac_addresses = set(input_file.read().splitlines())

# Write the unique MAC addresses to the output file
with open(path5, 'w') as output_file:
    output_file.write('\n'.join(mac_addresses))

input_file  =  path5
output_file = (path0 + "\scraping.xlsx")

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

with open(input_file, 'r') as data:
    reader = csv.reader(data, delimiter='\t')
    for row in reader:
        ws.append(row)

wb.save(output_file)

# Define the paths to the Excel files
file1_path = output_file
file2_path = read_file_path2("C:\Attendance\Code\_Important.txt")
file4_path = read_file_path1("C:\Attendance\Code\_Important.txt")
file3_path = (file4_path + "\Out.xlsx")

# Load both files using openpyxl
file1 = openpyxl.load_workbook(file1_path)
file2 = openpyxl.load_workbook(file2_path)

# Get the first sheet of each file
sheet1 = file1.active
sheet2 = file2.active

# Iterate over the cells in the first column of the first file
for cell in sheet1['A']:
    # Get the value of the cell in the first file
    value = cell.value
    # Iterate over the cells in the second column of the second file
    for cell2 in sheet2['B']:
        # Get the value of the cell in the second file
        value2 = cell2.value
        # If the values match, write '1' in the cell to the right of the matching cell in the second file
        if value == value2:
            sheet2.cell(row=cell2.row, column=cell2.column+1).value = 1

# Save the changes to the second file
file2.save(file3_path)

time.sleep(0.3)
#To change the excel with date
# Load the Excel file
workbook = openpyxl.load_workbook(file3_path)
worksheet = workbook.active

worksheet.delete_cols(2)

# Get today's date and format it as 'day-month'
today = datetime.date.today().strftime('%d-%m')

# Set the output file path
output_path = (file4_path + "\Attendance")

# Create the output file name
output_file = f'_{today}.xlsx'

# Save the modified workbook to a new Excel file
workbook.save(output_path + output_file)



